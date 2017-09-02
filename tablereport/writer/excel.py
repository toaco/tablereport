import math

import six
from openpyxl.styles import Alignment, Side, Border, Font, PatternFill
from openpyxl.utils import get_column_letter


class WorkSheetWriter(object):
    @staticmethod
    def write(worksheet, table, position):
        row_height = [None] * table.height
        col_width = [None] * table.width
        x, y = position[0] + 1, position[1] + 1

        for row_num in range(table.height):
            for col_num in range(table.width):
                cell = table[row_num][col_num]

                if cell is None:
                    continue

                excel_x = x + row_num
                excel_y = y + col_num

                if any([cell.height > 1, cell.width > 1]):
                    worksheet.merge_cells(
                        start_row=excel_x,
                        end_row=excel_x + cell.height - 1,
                        start_column=excel_y,
                        end_column=excel_y + cell.width - 1)

                excel_cell = worksheet.cell(row=excel_x, column=excel_y,
                                            value=cell.value)

                if cell.style is None:
                    continue

                font_weight = cell.style.get('font_weight')
                font_size = cell.style.get('font_size')
                if font_weight is not None or font_size is not None:
                    font = Font(size=font_size, bold=font_weight == 'blod')
                    excel_cell.font = font

                vertical_align = cell.style.get('vertical_align')
                horizontal_align = cell.style.get('horizontal_align')
                if vertical_align is not None or horizontal_align is not None:
                    align = Alignment(horizontal=horizontal_align,
                                      vertical=vertical_align)
                    excel_cell.alignment = align

                background_color = cell.style.get('background_color')
                if background_color is not None:
                    fill = PatternFill(start_color=background_color,
                                       end_color=background_color,
                                       fill_type='darkDown')
                    excel_cell.fill = fill

                side = Side(border_style='thin', color="fff0f0f0")
                border = Border(
                    left=excel_cell.border.left,
                    right=excel_cell.border.right,
                    top=excel_cell.border.top,
                    bottom=excel_cell.border.bottom
                )
                border.left = side
                border.right = side
                border.top = side
                border.bottom = side
                excel_cell.border = border

                if all([cell.height == 1, cell.width == 1]):
                    font_size = font_size or 11
                    width = cell.style.get('width')
                    if width is not None:
                        if width == 'auto':
                            width = (len(
                                six.text_type(cell.value).encode('utf-8'))
                                     + len(six.text_type(cell.value))
                                     ) / 2 * math.ceil(font_size / 11.0)
                        col_width[col_num] = max(width, col_width[col_num],
                                                 key=lambda v: v or 0)

                    height = cell.style.get('height')
                    if height is not None:
                        if height == 'auto':
                            height = math.ceil(font_size * 1.5)
                        row_height[row_num] = max(height, row_height[row_num],
                                                  key=lambda v: v or 0)
                else:
                    height = cell.style.get('height')
                    if height is not None:
                        if height == 'auto':
                            height = math.ceil(font_size * 1.5)
                        row_height[row_num] = max(height, row_height[row_num],
                                                  key=lambda v: v or 0)

        for i, value in enumerate(row_height):
            if value is None:
                pass
            else:
                worksheet.row_dimensions[position[0] + i + 1].height = value
        for i, value in enumerate(col_width):
            if value is None:
                pass
            else:
                column_letter = get_column_letter(position[1] + i + 1)
                worksheet.column_dimensions[column_letter].width = value
