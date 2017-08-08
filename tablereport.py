#!/usr/bin/python
# -*- coding: utf-8 -*-
from __future__ import unicode_literals

import math

from openpyxl.styles import Alignment, Side, Border, Font, PatternFill
from openpyxl.utils import get_column_letter


class Areas(list):
    def __init__(self, areas=None):
        if areas is None:
            areas = []
        super(Areas, self).__init__(areas)

    def merge(self, style=None):
        for area in self:
            area.merge(style)

    def add_summary(self, text_span, text, location, label_style=None,
                    value_style=None):
        for area in self:
            area.add_summary(text_span, text, location, label_style,
                             value_style)

    def one(self):
        """assert Areas contain only one Area and return it"""
        assert len(self) == 1
        return self[0]


class Area(object):
    def __init__(self, table, width, height, position, style=None):
        self.table = table
        self.table.areas.append(self)

        self.width = width
        self.height = height

        self._x, self._y = position

        self.style = style

    @property
    def position(self):
        return self._x, self._y

    @position.setter
    def position(self, value):
        self._x, self._y = value

    @property
    def data(self):
        rows = []
        for row_num in xrange(self.height):
            position = self._x + row_num, self._y
            row = Row(self.table, position, self.width)
            rows.append(row)
        return rows

    def merge(self, style=None):
        cell = self.data[0][0]

        for row in xrange(len(self.data)):
            for col in xrange(len(self.data[0])):
                self.data[row][col] = None
        cell.height = len(self.data)
        if style is not None:
            cell.style = style

        self.data[0][0] = cell

    def add_summary(self, text_span, text, location, label_style=None,
                    value_style=None):
        if location == 'left':
            cell = self.data[0][0]
            cell.height += 1
            new_col_num = self._add_row_at_bottom(
                label_style, text, text_span, value_style, self.width)
        elif location == 'down':
            new_col_num = self._add_row_at_bottom(
                label_style, text, text_span, value_style, offset=0)
        else:
            raise ValueError
        self._update_existed_areas(new_col_num)

        self.table.height += 1

    def _update_existed_areas(self, new_col_num):
        for area in self.table.areas:
            x, y = area.position
            if new_col_num > x + area.height:
                continue
            elif x + area.height >= new_col_num > x:
                area.height += 1
            else:
                x += 1
                area.position = x, y

    def _add_row_at_bottom(self, label_style, text, text_span, value_style,
                           offset=0):
        new_col_num = self._x + self.height
        self.table.data.insert(new_col_num,
                               [None] * self.table.width)
        appended_row = self.table[new_col_num]

        # set summary cell
        if text_span != 0:
            appended_row[self._y + offset] = Cell(text, width=text_span)
            if label_style is not None:
                appended_row[self._y + offset].style = label_style

        # summarize columns need to be summarized
        for col_num in xrange(self._y + offset + text_span,
                              self.table.width):
            total = 0
            for row_num in xrange(self._x, self._x + self.height):
                if row_num in self.table.total_row_nums:
                    continue
                total += self.table[row_num][col_num].value
            appended_row[col_num] = Cell(total)
            if value_style is not None:
                appended_row[col_num].style = value_style
            else:
                appended_row[col_num].style = self.table.style
        self.table.total_row_nums.add(new_col_num)
        return new_col_num

    def select(self, selector):
        # select an area in self
        area = selector.select(self)
        return area

    def __getitem__(self, item):
        return self.table.data[item]

    def __setitem__(self, key, value):
        self.table.data[key] = value


class Table(object):
    def __init__(self, headers=None, body=None, style=None):

        if headers is None:
            headers = []

        if body is None:
            body = []

        self.headers = headers
        self.body = body

        self.table = self.headers + self.body
        data = self.table

        for row_num in xrange(len(data)):
            for col_num in xrange(len(data[0])):
                cell = data[row_num][col_num]
                if cell is not None:
                    if isinstance(cell, tuple):
                        data[row_num][col_num] = Cell(cell[0], style=cell[1])
                    else:
                        data[row_num][col_num] = Cell(data[row_num][col_num],
                                                      style=style)
                    self._auto_merge(data, row_num, col_num)
        self._data = data
        self.areas = []
        self.total_row_nums = set()
        try:
            width = len(self._data[0])
        except IndexError:
            width = 0

        self.width = width
        self.height = len(self._data)
        self.style = style

    def __getitem__(self, item):
        return self._data[item]

    def __setitem__(self, key, value):
        self._data[key] = value

    @property
    def data(self):
        return self._data

    @staticmethod
    def _auto_merge(data, row_num, col_num):
        # todo: range judge
        for i in xrange(row_num + 1, len(data)):
            if data[i][col_num] is None:
                data[row_num][col_num].height += 1
            else:
                break

        for j in xrange(col_num + 1, len(data[0])):
            if data[row_num][j] is None:
                data[row_num][col_num].width += 1
            else:
                break

    def select(self, selector):
        # select an area in self
        sub_area = Area(table=self, width=self.width, height=len(self.body),
                        position=(len(self.headers), 0))
        areas = selector.select(sub_area)
        return areas

    def add_summary(self, text_span, text, location, label_style=None,
                    value_style=None):
        # todo width and height should auto change
        body = Area(table=self, width=self.width,
                    height=len(self.data) - len(self.headers),
                    position=(len(self.headers), 0))
        body.add_summary(text_span, text, location, label_style, value_style)


class Row(object):
    def __init__(self, table, position, width):
        self.table = table
        self.x, self.y = position
        self.width = width

    def __getitem__(self, col):
        assert col < self.width
        return self.table[self.x][self.y + col]

    def __setitem__(self, col, value):
        assert col < self.width
        self.table[self.x][self.y + col] = value

    def __eq__(self, iterable):
        return all(self[i] == iterable[i] for i in xrange(self.width))

    def __len__(self):
        return self.width

    def __repr__(self):
        return str([self[i] for i in xrange(self.width)])


class Cell(object):
    def __init__(self, value, style=None, width=1, height=1):
        self.value = value
        self.width = width
        self.height = height
        self.style = style

    def __eq__(self, other):
        if type(other) == Cell:
            return self.__dict__ == other.__dict__
        else:
            assert type(self.value) == type(other)
            return self.value == other

    def __repr__(self):
        return ('Cell(value="{}", style={}, width={}, height={})'
                .format(repr(self.value), self.style, self.width, self.height)
                .encode('utf-8'))

    __str__ = __repr__


class ColumnSelector:
    def __init__(self, column, group):
        self.column = column
        self.group = group

    def select(self, area):
        assert self.column <= area.width
        x, y = area.position

        y += self.column - 1
        area = Area(table=area.table, width=1, height=area.height,
                    position=(x, y))

        areas = Areas()
        if not self.group:
            areas.append(area)
        else:
            start_value = area.data[0][0]
            start_index = 0
            for row_num in xrange(1, len(area.data)):
                if area.data[row_num][0] == start_value:
                    continue
                else:
                    sub_area = Area(table=area.table, width=1,
                                    height=row_num - start_index,
                                    position=(x + start_index, y))
                    start_value = area.data[row_num][0]
                    start_index = row_num
                    areas.append(sub_area)
            else:
                sub_area = Area(table=area.table, width=1,
                                height=len(area.data) - start_index,
                                position=(x + start_index, y))
                areas.append(sub_area)
        return areas


class ExcelWriter(object):
    @staticmethod
    def write(worksheet, table, position):
        row_height = [None] * table.height
        col_width = [None] * table.width
        x, y = position[0] + 1, position[1] + 1

        for row_num in xrange(table.height):
            for col_num in xrange(table.width):
                cell = table[row_num][col_num]

                if cell is None:
                    continue

                excel_x = x + row_num
                excel_y = y + col_num

                if any([cell.height > 1, cell.width > 1]):
                    worksheet.merge_cells(
                        start_row=excel_x,
                        end_row=excel_x + cell.height - 1,
                        start_column=excel_y,
                        end_column=excel_y + cell.width - 1)

                excel_cell = worksheet.cell(row=excel_x, column=excel_y,
                                            value=cell.value)

                if cell.style is None:
                    continue

                font_weight = cell.style.get('font_weight')
                font_size = cell.style.get('font_size')
                if font_weight is not None or font_size is not None:
                    font = Font(size=font_size, bold=font_weight == 'blod')
                    excel_cell.font = font

                vertical_align = cell.style.get('vertical_align')
                horizontal_align = cell.style.get('horizontal_align')
                if vertical_align is not None or horizontal_align is not None:
                    align = Alignment(horizontal=horizontal_align,
                                      vertical=vertical_align)
                    excel_cell.alignment = align

                background_color = cell.style.get('background_color')
                if background_color is not None:
                    fill = PatternFill(start_color=background_color,
                                       end_color=background_color,
                                       fill_type='darkDown')
                    excel_cell.fill = fill

                side = Side(border_style='thin', color="fff0f0f0")
                border = Border(
                    left=excel_cell.border.left,
                    right=excel_cell.border.right,
                    top=excel_cell.border.top,
                    bottom=excel_cell.border.bottom
                )
                border.left = side
                border.right = side
                border.top = side
                border.bottom = side
                excel_cell.border = border

                if all([cell.height == 1, cell.width == 1]):
                    font_size = font_size or 11
                    width = cell.style.get('width')
                    if width is not None:
                        if width == 'auto':
                            width = (len(
                                unicode(cell.value).encode('utf-8')) + len(
                                unicode(cell.value))) / 2 * math.ceil(
                                font_size / 11.0)
                        col_width[col_num] = max(width, col_width[col_num])

                    height = cell.style.get('height')
                    if height is not None:
                        if height == 'auto':
                            height = math.ceil(font_size * 1.5)
                        row_height[row_num] = max(height, row_height[row_num])
                else:
                    height = cell.style.get('height')
                    if height is not None:
                        if height == 'auto':
                            height = math.ceil(font_size * 1.5)
                        row_height[row_num] = max(height, row_height[row_num])

        for i, value in enumerate(row_height):
            if value is None:
                pass
            else:
                worksheet.row_dimensions[i + 1].height = value
        for i, value in enumerate(col_width):
            if value is None:
                pass
            else:
                column_letter = get_column_letter(i + 1)
                worksheet.column_dimensions[column_letter].width = value


class Style(object):
    """
    Style and style check is here
    """

    def __new__(cls, dict_1=None, extend=None):
        if dict_1 is None:
            dict_1 = {}
        else:
            assert isinstance(dict_1, dict)

        if extend is not None:
            assert isinstance(extend, dict)
            extend = extend.copy()
            extend.update(dict_1)
            return extend
        return dict_1
